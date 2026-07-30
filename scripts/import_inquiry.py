"""
返送された質疑書(Excel)の回答を inquiry.md(Markdown) に書き戻すスクリプト。

`generate_inquiry.py` の逆方向。質疑事項の正本は Markdown なので、
客先が記入して返送したExcelの回答欄・ヘッダー欄をMarkdownに取り込む。

使い方:
    python scripts/import_inquiry.py <project_id> <返送されたxlsx> [--dry-run] [--overwrite]

例:
    python scripts/import_inquiry.py 20260614-001 ~/Downloads/inquiry_20260614-001.xlsx --dry-run
    python scripts/import_inquiry.py 20260614-001 ~/Downloads/inquiry_20260614-001.xlsx

動作:
    - No列が Qn のセルを照合し、Excelの回答欄をMarkdownの回答欄に書き込む
    - **Excelの回答が空欄の質問は触らない**（Markdown側の記入を消さない）
    - Markdown側に既に回答がある質問は既定でスキップし、警告を出す。
      上書きする場合は --overwrite を付ける
    - ヘッダー欄（宛先・設備名称・件名・回答期限・送付者・連絡先）も取り込む。
      Markdown側がプレースホルダ `{...}` か空欄の場合のみ（既定）
    - セル内改行は `<br>`、`|` は `\\|` にエスケープしてMarkdownの表を壊さない
    - 変更内容を一覧表示する。--dry-run なら書き込まずに表示のみ

Excelに無いQ番号・Markdownに無いQ番号は警告として報告する（質問を追加・削除した
状態のExcelが返ってきた場合に気付けるようにするため）。
"""

import sys
import os
import re
from openpyxl import load_workbook

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJECTS_DIR = os.path.join(REPO_ROOT, "database", "projects")

HEADER_FIELDS = ["宛先", "設備名称", "件名", "送付日", "回答期限", "送付者", "連絡先"]


def to_markdown_cell(value):
    """Excelのセル値をMarkdownの表セルに入れられる文字列にする"""
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("|", "\\|")
    text = re.sub(r"\r\n|\r|\n", "<br>", text)
    return text


def has_placeholder(text):
    return bool(re.search(r"\{[^}]*\}", text))


def read_xlsx(path):
    """返送されたExcelから回答とヘッダーを読む"""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    answers, header = {}, {}
    for row in range(1, ws.max_row + 1):
        col_a = ws.cell(row, 1).value
        if col_a is None:
            continue
        key = str(col_a).strip()

        if re.fullmatch(r"[Qq]\d+", key):
            answers[key.upper()] = to_markdown_cell(ws.cell(row, 3).value)
        elif key in HEADER_FIELDS:
            header[key] = to_markdown_cell(ws.cell(row, 2).value)

    return answers, header


def apply_to_markdown(md_path, answers, header, overwrite=False):
    """
    Markdownに反映する。戻り値は (新しい本文, 変更一覧, 警告一覧)。
    変更一覧の各要素は (対象, 変更前, 変更後)。
    """
    lines = open(md_path, encoding="utf-8").read().splitlines(True)
    changes, warnings = [], []
    in_questions = False
    seen = set()
    out = []

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## "):
            in_questions = (stripped == "## 質疑事項")

        m = re.match(r"^\| (Q\d+) \|", line)
        if in_questions and m:
            no = m.group(1)
            seen.add(no)
            cells = line.rstrip("\n").split("|")
            if len(cells) != 6:
                warnings.append(f"{no}: 4列の表ではないため触らない（{len(cells)-2}列）")
                out.append(line)
                continue
            current = cells[3].strip()
            new = answers.get(no)
            if new is None:
                warnings.append(f"{no}: Excel側に見つからない（質問を追加した？）")
            elif new == "":
                pass  # Excelが空欄 → 触らない
            elif new == current:
                pass  # 変化なし
            elif current and not overwrite:
                warnings.append(
                    f"{no}: Markdown側に既に回答があるためスキップ"
                    f"（上書きするなら --overwrite）\n"
                    f"      Markdown: {current}\n"
                    f"      Excel   : {new}"
                )
            else:
                cells[3] = f" {new} "
                line = "|".join(cells) + "\n"
                changes.append((no, current, new))
            out.append(line)
            continue

        m = re.match(r"^\| (" + "|".join(HEADER_FIELDS) + r") \|", line)
        if m and m.group(1) in header:
            field = m.group(1)
            cells = line.rstrip("\n").split("|")
            current = cells[2].strip()
            new = header[field]
            fillable = (current == "" or has_placeholder(current))
            if new and new != current and (fillable or overwrite):
                cells[2] = f" {new} "
                line = "|".join(cells) + "\n"
                changes.append((field, current, new))
            elif new and new != current:
                warnings.append(
                    f"{field}: Markdown側に値があるためスキップ"
                    f"（上書きするなら --overwrite）\n"
                    f"      Markdown: {current}\n"
                    f"      Excel   : {new}"
                )
            out.append(line)
            continue

        out.append(line)

    for no in sorted(set(answers) - seen, key=lambda s: int(s[1:])):
        if answers[no]:
            warnings.append(f"{no}: Markdown側に見つからない（回答: {answers[no][:40]}）")

    return "".join(out), changes, warnings


def main(argv):
    args = [a for a in argv if not a.startswith("--")]
    flags = {a for a in argv if a.startswith("--")}
    unknown = flags - {"--dry-run", "--overwrite"}
    if unknown:
        print(f"エラー: 不明なオプション: {', '.join(sorted(unknown))}", file=sys.stderr)
        return 1
    if len(args) < 2:
        print(__doc__.strip(), file=sys.stderr)
        return 1

    project_id, xlsx = args[0], args[1]
    md_path = os.path.join(PROJECTS_DIR, project_id, "inquiry.md")

    for path, label in ((xlsx, "Excel"), (md_path, "Markdown")):
        if not os.path.exists(path):
            print(f"エラー: {label}が見つかりません: {path}", file=sys.stderr)
            return 1

    answers, header = read_xlsx(xlsx)
    if not answers:
        print(
            f"エラー: 回答を読み取れませんでした: {xlsx}\n"
            "A列に Q1・Q2… があり、C列が回答欄である必要があります"
            "（generate_inquiry.py が出力した形式）",
            file=sys.stderr,
        )
        return 1

    body, changes, warnings = apply_to_markdown(
        md_path, answers, header, overwrite=("--overwrite" in flags)
    )

    print(f"Excel   : {xlsx}")
    print(f"Markdown: {md_path}")
    print(f"読み取り: 回答{len(answers)}問・ヘッダー{len(header)}項目\n")

    if changes:
        print(f"取り込む変更 {len(changes)}件:")
        for target, before, after in changes:
            print(f"  {target}: {before or '(空欄)'} → {after}")
    else:
        print("取り込む変更はありません。")

    if warnings:
        print(f"\n警告 {len(warnings)}件:")
        for w in warnings:
            print(f"  {w}")

    if "--dry-run" in flags:
        print("\n--dry-run のため書き込みませんでした。")
        return 0

    if changes:
        open(md_path, "w", encoding="utf-8").write(body)
        print(f"\n書き込み完了: {md_path}")
        print(f"Excelを作り直す: python scripts/generate_inquiry.py {project_id}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
