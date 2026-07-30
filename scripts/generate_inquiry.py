"""
質疑書(Markdownドラフト)をExcel形式に変換するスクリプト。

質疑事項の正本は `inquiry.md`(Markdown)である。本スクリプトは内容を持たず、
Markdownを読んでExcelに変換するだけの変換器として動作する。
質問の追加・修正はMarkdown側で行うこと(このファイルを編集しない)。

読み込み順:
    1. database/projects/{project_id}/inquiry.md   ← 案件のドラフト(Agent01作成)
    2. database/projects/_template/inquiry.md      ← 1が無い場合のひな形

使い方:
    python scripts/generate_inquiry.py <project_id> [output_dir]
    python scripts/generate_inquiry.py <project_id> [output_dir] --source <path.md>

例:
    python scripts/generate_inquiry.py 20260614-001
        → database/projects/20260614-001/inquiry.md を読み
          database/projects/20260614-001/inquiry_20260614-001.xlsx を出力

    python scripts/generate_inquiry.py 20260614-001 /tmp
    python scripts/generate_inquiry.py template . --source path/to/inquiry.md
"""

import sys
import os
import re
import unicodedata
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJECTS_DIR = os.path.join(REPO_ROOT, "database", "projects")
TEMPLATE_MD = os.path.join(PROJECTS_DIR, "_template", "inquiry.md")

# Excelヘッダーに載せる項目と並び順。Markdownのヘッダー表に無い項目は空欄で出す
HEADER_FIELDS = ["宛先", "設備名称", "件名", "送付日", "回答期限", "送付者", "連絡先"]

# ウインドウ枠の固定を行う上限（表見出しがこれより下にある場合は固定しない）。
# 固定すると見出しより上の行もすべて画面上部に貼り付くため、上が長い文書では逆に見づらい。
MAX_FREEZE_ROWS = 12

# 列幅（openpyxlの単位＝半角文字1文字ぶん。全角文字は2を占める）
# A列はNo欄だが、ヘッダー欄のラベル（「設備名称」「回答期限」= 全角4文字 = 幅8）も
# 入るため、余白を含めて収まる10にしている
COL_WIDTH = {"A": 10, "B": 50, "C": 50, "D": 50}

# 1行あたりの行高（ポイント）。メイリオ10ptで折り返し1行に必要な高さ
LINE_HEIGHT_PT = 16

# --- スタイル定義 ---
COLOR_HEADER_BG  = "1F4E79"  # 濃紺
COLOR_SECTION_BG = "BDD7EE"  # 薄青
COLOR_NOTE_BG    = "FFF2CC"  # 薄黄
COLOR_INPUT_BG   = "F2F9FF"  # 記入欄の水色
COLOR_WHITE      = "FFFFFF"
COLOR_BLACK      = "000000"

FONT_NAME = "メイリオ"


# ===========================================================================
# Markdown 解析
# ===========================================================================

def _split_row(line):
    """Markdownテーブルの1行をセルのリストに分解する"""
    line = line.strip()
    if not line.startswith("|"):
        return None
    # 末尾の | を落としてから分割(\| はセル内のリテラル)
    body = line[1:]
    if body.endswith("|"):
        body = body[:-1]
    cells = re.split(r"(?<!\\)\|", body)
    return [_clean_cell(c) for c in cells]


def _clean_cell(text):
    """セル内のMarkdown記法をExcel向けのプレーンテキストに直す"""
    text = text.strip()
    text = text.replace("\\|", "|")
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)  # 太字
    text = re.sub(r"`(.+?)`", r"\1", text)        # コード
    return text


def _is_separator(cells):
    """`|---|---|` の区切り行かどうか"""
    return bool(cells) and all(re.fullmatch(r":?-{2,}:?", c) for c in cells if c != "")


def parse_inquiry_md(path):
    """
    inquiry.md を解析して (header, sections) を返す。

    header   : {"宛先": "...", ...}
    sections : [(節タイトル, [(No, 確認事項, 回答, 備考), ...]), ...]
    """
    with open(path, encoding="utf-8") as f:
        lines = f.read().splitlines()

    header = {}
    sections = []
    body = []                # 「## 本文」の内容: ("heading"|"para"|"row", ...)
    current_section = None   # (title, rows) — 質疑事項の節を読んでいる間だけ非None
    in_header_table = False
    in_questions = False     # 「## 質疑事項」以降か
    in_body = False          # 「## 本文」以降か

    for line in lines:
        stripped = line.strip()

        if stripped.startswith("## "):
            title = stripped[3:].strip()
            in_header_table = (title == "ヘッダー")
            in_questions = (title == "質疑事項")
            in_body = (title == "本文")
            current_section = None
            continue

        if stripped.startswith("### "):
            # 「### 【設備・装置について】」→「設備・装置について」
            title = stripped[4:].strip().strip("【】")
            if in_questions:
                current_section = (title, [])
                sections.append(current_section)
            elif in_body:
                body.append(("heading", title))
            continue

        cells = _split_row(stripped)

        # 本文は表でない行も拾うので、表の判定より先に処理する
        if in_body:
            if cells and not _is_separator(cells):
                if len(cells) >= 2 and cells[0] not in ("項目", "内容"):
                    body.append(("row", _clean_cell(cells[0]), _clean_cell(cells[1])))
            elif (
                not cells
                and stripped
                and not stripped.startswith(">")
                and not re.fullmatch(r"-{3,}|\*{3,}|_{3,}", stripped)  # 水平線
            ):
                body.append(("para", _clean_cell(stripped)))
            continue

        if not cells or _is_separator(cells):
            continue

        if in_header_table and len(cells) >= 2:
            key, value = cells[0], cells[1]
            if key and key not in ("項目", "内容"):
                header[key] = value
            continue

        if current_section is not None:
            # No列がQ番号の行のみ質問として扱う(見出し行を自然に除外できる)
            if not re.fullmatch(r"[Qq]\d+", cells[0]):
                continue
            no      = cells[0].upper()
            quest   = cells[1] if len(cells) > 1 else ""
            answer  = cells[2] if len(cells) > 2 else ""
            note    = cells[3] if len(cells) > 3 else ""
            current_section[1].append((no, quest, answer, note))

    # 質問が1件も無い節は落とす(表を持たない節見出しが混ざった場合)
    sections = [(t, rows) for t, rows in sections if rows]
    return header, sections, body


def format_question(text):
    """
    末尾の選択肢カッコを改行して読みやすくする。
    「...ありますか？（AC200V / ...）」→「...ありますか？\n（AC200V / ...）」
    """
    return re.sub(r"([。？?])\s*（", r"\1\n（", text)


def has_placeholder(text):
    """`{客先名}` のような未記入プレースホルダを含むか"""
    return bool(re.search(r"\{[^}]*\}", text))


# ===========================================================================
# Excel 出力
# ===========================================================================

def thin_border():
    s = Side(style="thin", color=COLOR_BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def display_width(text):
    """
    文字列の表示幅を返す（全角=2、半角=1）。
    列幅と同じ単位にするため。**全角を1と数えると行数を半分に見誤り、
    行高が足りずに文字が見切れる。**
    """
    width = 0
    for ch in str(text):
        width += 2 if unicodedata.east_asian_width(ch) in ("F", "W", "A") else 1
    return width


def span_width(*cols):
    """結合したセルの合計幅（COL_WIDTH の単位）"""
    return sum(COL_WIDTH[c] for c in cols)


def wrapped_lines(text, width_units):
    """折り返し後の行数。明示的な改行も数える"""
    if not text:
        return 1
    usable = max(width_units - 2, 4)   # 左右の余白ぶんを引く
    lines = 0
    for segment in str(text).split("\n"):
        lines += max(1, -(-display_width(segment) // usable))  # 切り上げ除算
    return lines


def row_height(*measured, minimum=1):
    """
    (テキスト, 幅) の組から必要な行高（pt）を求める。
    最も行数が多い列に合わせる。
    """
    lines = max([wrapped_lines(t, w) for t, w in measured] + [minimum])
    return lines * LINE_HEIGHT_PT


def make_header(ws, project_id, header, title="質 疑 書", preamble=None):
    """タイトル・ヘッダー部を作成する"""
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"{title}　－　{project_id}"
    c.font = Font(name=FONT_NAME, size=16, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    row = 2
    for label in HEADER_FIELDS:
        value = header.get(label, "")
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = Font(name=FONT_NAME, size=10, bold=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].border = thin_border()

        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"].value = value
        ws[f"B{row}"].font = Font(name=FONT_NAME, size=10)
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{row}"].border = thin_border()
        ws[f"B{row}"].alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        # 未記入(空欄またはプレースホルダ)は水色にして手入力箇所を示す
        if value == "" or has_placeholder(value):
            ws[f"B{row}"].fill = PatternFill("solid", fgColor=COLOR_INPUT_BG)
        ws.row_dimensions[row].height = row_height(
            (value, span_width("B", "C", "D"))
        )
        row += 1

    # 前文
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = preamble or (
        "下記の事項についてご確認をお願いいたします。"
        "ご多忙のところ恐れ入りますが、回答期限までにご回答いただけますと幸いです。"
    )
    ws[f"A{row}"].font = Font(name=FONT_NAME, size=10)
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = row_height(
        (ws[f"A{row}"].value, span_width("A", "B", "C", "D"))
    )
    return row + 1


def make_body(ws, row, body):
    """「## 本文」の説明ブロックを書く（承認依頼のように説明が要る文書向け）"""
    for item in body:
        kind = item[0]
        if kind == "heading":
            ws.merge_cells(f"A{row}:D{row}")
            c = ws[f"A{row}"]
            c.value = f"■ {item[1]}"
            c.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_HEADER_BG)
            c.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()
            ws.row_dimensions[row].height = 22
        elif kind == "para":
            ws.merge_cells(f"A{row}:D{row}")
            c = ws[f"A{row}"]
            c.value = item[1]
            c.font = Font(name=FONT_NAME, size=10)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = row_height(
                (item[1], span_width("A", "B", "C", "D"))
            )
        else:  # ("row", ラベル, 値)
            label, value = item[1], item[2]
            # ラベルはA列(幅8=全角4文字)では狭すぎるためA:Bに結合する
            ws.merge_cells(f"A{row}:B{row}")
            ws[f"A{row}"].value = label
            ws[f"A{row}"].font = Font(name=FONT_NAME, size=10, bold=True)
            ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
            ws[f"A{row}"].alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            ws[f"A{row}"].border = thin_border()

            ws.merge_cells(f"C{row}:D{row}")
            ws[f"C{row}"].value = value
            ws[f"C{row}"].font = Font(name=FONT_NAME, size=10)
            ws[f"C{row}"].alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            ws[f"C{row}"].border = thin_border()
            ws.row_dimensions[row].height = row_height(
                (label, span_width("A", "B")),
                (value, span_width("C", "D")),
            )
        row += 1

    if body:
        row += 1  # 質疑表との間を1行あける
    return row


def make_table_header(ws, row):
    labels = {"A": "No", "B": "確認事項", "C": "回　答　欄", "D": "備考"}
    for col, label in labels.items():
        c = ws[f"{col}{row}"]
        c.value = label
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 20
    return row + 1


def make_section(ws, row, section_title):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws[f"A{row}"]
    c.value = f"■ {section_title}"
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_HEADER_BG)
    c.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    ws.row_dimensions[row].height = 20
    return row + 1


def make_question(ws, row, no, question, answer, note):
    # No列
    c = ws[f"A{row}"]
    c.value = no
    c.font = Font(name=FONT_NAME, size=10, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    c.border = thin_border()

    # 確認事項列
    c = ws[f"B{row}"]
    c.value = question
    c.font = Font(name=FONT_NAME, size=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border()

    # 回答欄(Markdown側に回答が書かれていれば引き継ぐ)
    c = ws[f"C{row}"]
    c.value = answer
    c.font = Font(name=FONT_NAME, size=10)
    if answer == "":
        c.fill = PatternFill("solid", fgColor=COLOR_INPUT_BG)
    c.border = thin_border()
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 備考
    c = ws[f"D{row}"]
    c.value = note
    c.font = Font(name=FONT_NAME, size=9, color="595959")
    if note:
        c.fill = PatternFill("solid", fgColor=COLOR_NOTE_BG)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border()

    # 行高さ: 改行数とB列幅(50文字換算)からの折り返し推定の大きい方
    ws.row_dimensions[row].height = row_height(
        (question, COL_WIDTH["B"]),
        (answer, COL_WIDTH["C"]),
        (note, COL_WIDTH["D"]),
        minimum=2,   # 回答を書き込む余地として最低2行ぶん確保する
    )
    return row + 1


# ===========================================================================
# エントリポイント
# ===========================================================================

# 文書種別ごとの設定。--kind で切り替える
KINDS = {
    "inquiry": {
        "file": "inquiry.md",
        "title": "質 疑 書",
        "sheet": "質疑書",
        "preamble": None,   # 既定の前文を使う
    },
    "approval": {
        "file": "safety_approval.md",
        "title": "安全対策 ご確認のお願い",
        "sheet": "安全対策確認",
        "preamble": (
            "弊社にて安全対策を検討いたしましたので、下記のとおりご確認をお願いいたします。"
            "内容にご要望・相違がございましたらご指摘ください。"
        ),
    },
}


def resolve_source(project_id, source=None, kind="inquiry"):
    """読み込むMarkdownのパスを決める"""
    if source:
        if not os.path.exists(source):
            raise FileNotFoundError(f"指定されたMarkdownが見つかりません: {source}")
        return source

    filename = KINDS[kind]["file"]
    project_md = os.path.join(PROJECTS_DIR, project_id, filename)
    if os.path.exists(project_md):
        return project_md
    # ひな形へのフォールバックは質疑書のみ（承認依頼は案件ごとの内容なのでひな形がない）
    if kind == "inquiry" and os.path.exists(TEMPLATE_MD):
        return TEMPLATE_MD
    raise FileNotFoundError(f"Markdownが見つかりません: {project_md}")


def default_output_dir(project_id):
    """案件フォルダがあればその中、無ければカレント"""
    project_dir = os.path.join(PROJECTS_DIR, project_id)
    return project_dir if os.path.isdir(project_dir) else "."


def generate(project_id, output_dir=None, source=None, kind="inquiry"):
    conf = KINDS[kind]
    src = resolve_source(project_id, source, kind)
    header, sections, body = parse_inquiry_md(src)

    if not sections:
        raise ValueError(
            f"質疑事項を読み取れませんでした: {src}\n"
            "「## 質疑事項」の下に「### 【節名】」と "
            "|No|確認事項|回答欄|備考| の表が必要です。"
        )

    # {project_id} と 送付日 だけは機械的に埋める。他のプレースホルダは
    # 人間が判断する項目なので残したまま水色表示にする
    header = {
        k: v.replace("{project_id}", project_id) for k, v in header.items()
    }
    if not header.get("送付日") or has_placeholder(header.get("送付日", "")):
        header["送付日"] = date.today().strftime("%Y-%m-%d")

    wb = Workbook()
    ws = wb.active
    ws.title = conf["sheet"]

    # 列幅は COL_WIDTH が正本。行高の計算にも同じ値を使う（二重に持たない）
    for col, width in COL_WIDTH.items():
        ws.column_dimensions[col].width = width

    row = make_header(ws, project_id, header, conf["title"], conf["preamble"])
    row = make_body(ws, row, body)
    row = make_table_header(ws, row)
    header_row = row - 1

    count = 0
    for section_title, questions in sections:
        row = make_section(ws, row, section_title)
        for no, question, answer, note in questions:
            row = make_question(ws, row, no, format_question(question), answer, note)
            count += 1

    # 印刷設定
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = f"{header_row}:{header_row}"  # 各ページに表見出しを繰り返す

    # ウインドウ枠の固定は表見出しの直下。ただしExcelは飛び飛びの行を固定できないため、
    # 見出しより上（ヘッダー欄・本文）もすべて固定されてしまう。上が長い文書で固定すると
    # 画面がほぼ埋まって中身を確認できなくなるので、固定行数に上限を設ける。
    if header_row <= MAX_FREEZE_ROWS:
        ws.freeze_panes = f"A{header_row + 1}"

    out_dir = output_dir or default_output_dir(project_id)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{kind}_{project_id}.xlsx")
    wb.save(out_path)

    print(f"読込元　: {src}")
    print(f"生成完了: {out_path}（{len(sections)}節 / {count}問）")
    return out_path


def main(argv):
    args, source, kind = [], None, "inquiry"
    i = 0
    while i < len(argv):
        if argv[i] in ("--source", "--kind"):
            if i + 1 >= len(argv):
                print(f"エラー: {argv[i]} の後に値を指定してください", file=sys.stderr)
                return 1
            if argv[i] == "--source":
                source = argv[i + 1]
            else:
                kind = argv[i + 1]
                if kind not in KINDS:
                    print(
                        f"エラー: --kind は {' / '.join(KINDS)} のいずれか（指定値: {kind}）",
                        file=sys.stderr,
                    )
                    return 1
            i += 2
            continue
        args.append(argv[i])
        i += 1

    project_id = args[0] if args else "template"
    output_dir = args[1] if len(args) > 1 else None

    try:
        generate(project_id, output_dir, source, kind)
    except (FileNotFoundError, ValueError) as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
